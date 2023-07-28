import os
import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
import warnings

warnings.filterwarnings("ignore", category=UserWarning)

# 文件夹路径
folder_path = "hbasepresentdata"

# 创建 hbasemap
hbasemap = {}

# 遍历文件夹中的每个文件
for filename in os.listdir(folder_path):
    file_path = os.path.join(folder_path, filename)

    if os.path.isfile(file_path) and filename.endswith(".xlsx"):
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active

        # 从第二行开始遍历第二列和第三列数据
        for row in sheet.iter_rows(min_row=2, min_col=2, max_col=3, values_only=True):
            key, value = row[0], row[1]

            hbasemap.setdefault(key, []).append(value)

# 读取 lzzkdata.xlsx 文件
lzzkdata_file = "lzzkdata.xlsx"
lzzkdata_workbook = openpyxl.load_workbook(lzzkdata_file)
lzzkdata_sheet = lzzkdata_workbook.active

# 创建 lzzkmap
lzzkmap = {}

# 从第二行开始遍历 B 列和 D 列数据
for row in lzzkdata_sheet.iter_rows(min_row=2, min_col=2, max_col=4, values_only=True):
    key, value = row[0], row[2]

    lzzkmap.setdefault(key, []).append(value)

# 扫描当前文件夹下的 output.xlsx
output_file = "output.xlsx"
output_workbook = openpyxl.load_workbook(output_file)
output_sheet = output_workbook.active

# 设置红色字体和背景颜色
font_red = Font(color="FF0000")
fill_red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
# 遍历 C 列和 E 列的值，并写入差集和填写 F 列
for row in output_sheet.iter_rows(min_row=1, min_col=0, max_col=7):
    c_value = row[2].value
    e_value = row[4].value
    g_cell = row[6]  # G 列的单元格
    f_cell = row[5]  # F 列的单元格

    if c_value in lzzkmap:
        lzzkmap_values = lzzkmap[c_value]
        difference = list(set(lzzkmap_values) - set(map(int, hbasemap.get(e_value, []))))

        # 写入差集至 G 列
        for value in difference:
            g_cell.value = value
            g_cell = output_sheet.cell(row=g_cell.row, column=g_cell.column + 1)

        # 填写 F 列
        if difference:
            f_cell.value = "有缺项: "
            f_cell.fill = fill_red  # 设置背景颜色
            f_cell.font = font_red  # 设置字体颜色
        else:
            f_cell.value = "无缺项"

        # 修改为填写在 F 列
        f_cell.offset(column=1).value = f_cell.offset(column=2).value
        f_cell.offset(column=2).value = None

# 保存修改后的文件
output_workbook.save("output_modified.xlsx")
